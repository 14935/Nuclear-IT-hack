# -*- coding: cp1251 -*-
import telebot
from gigachat import GigaChat
from similar_text import similar_text
import pandas as pd
from similar_text import similar_text
from dateutil import parser
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
import dateparser

def extract_relative_date_from_text(text):
    date = datetime(2024, 4, 3)
    # ���������� ���������� dateparser ��� ���������� ������������� ����
    extracted_date = dateparser.parse(text, settings={"RELATIVE_BASE": date})
    return extracted_date

def classify(text):
    input_text = text
    with open('prompt.txt', 'r',encoding="utf-8") as file:
        text = file.read()
    #print(text)
    with GigaChat(credentials='MmQyNzJmNzAtNTRhMi00ZTg5LTgxNGMtMWQ2OTJhNzg0MmFmOmU5MTY0YjlmLTVhODMtNGI2Yi1iZTU5LTJkYzM5ZjJiNjk3YQ==', verify_ssl_certs=False) as giga:
        #response_text = text + input_text
        response_text="������ ����� �������� ���� � ������������ ���� (�������� '1 �����', ��� '���������'), ������ ���� �� ������, ������ ��� ������� 03.04.2024 � ������ ������ ���� � ������� '<��.��.����>, �����: '"+input_text
        response = giga.chat(response_text)
        date1 = response.choices[0].message.content
    print(date1)
    candidate_labels = ['�.�������������', '������������', '������������. ��', '���������� ��', '��������������', '�����������. ��', '������� ������', '������ �����', '�������',
    '������� ���', '�������������', '���� �������� ��', '�����������', '����������', '��������� ����', '�����������', '��-� ��������.��', '���-��������',
    '���������������.', '�������������', '�������', '��������', '��������������', '�����������', '�������������', '���������� ��', '�������������',
    '�����������(���)', '��������', '����������', '����������� ��', '������', '��������', '�����', '����������', '������ �������', '������ ������',
    '���������', '���������', '����������', '�������� ���', '���������� ���', '��������� ���', '��. ���������', '������� ���', '����������', '����������-� ���',
    '����������', '������������', '������������', '������������', 'ٸ��������', '���������. ���', '��������� ��', '���������� ��', '������������',
    '����������� ��', '����', '���������������', '�������� ����', '����������', '���������� ��', '���������', '����������', '����������� ��',
    '��������������', '�������� ���� ��', '������������� ��', '������� ��', '��������� ��', '���������� ��', '������������', '����������� ��',
    '���� �������� ��', '�������� ��', '���������������.', '�����������', '������', '�. �����������', '���������-� ����', '������� ������',
    '������������', '�����������.����', '����������', '������������', '��������', '������������ ���', '����', '������������', '�������',
    '��-�� ���� ���', '�����������', '������������', '�����-����� ���', '�����������. ���', '����������� ���', '�����������', '��������� ��-�',
    '�������������', '�����������', '����� ��������', '���������', '�������', '��������', 'Ҹ���� ����', '�������', '��������������',
    '������', '��������� ��-�', '���������', '������������ ���', '������������� ��', '������������', '��������� ���', '�����-�����(�-�)',
    '��������� ����', '����������', '�����������', '����� 1905 ����', '�������', '������������', '����������� ����', '���������', '���������',
    '�����������', '���������', '���������', '��������', '��������', '���������', '����.-���������.', '�������������', '�����������', '���������� ���',
    '�������������', '������� �������', '���������', '����������', '�������', '������������', '��������', '�����������', '��������', '����������� ��-�',
    '���������������', '������������', '�����', '��������', '����������', '�������', '�����. �������', '�����������', '��������� ���', '��������', '�������',
    '�������������', '�������', '�����', '����. ��. ������', '��������', '��. ��. ������', '������', '����������������', '�-� ��. ��������',
    '���� ������ ���', '����������������', '��. ������������', '�-� ���. �������', '��. ���������', '��������� �����', '�������', '���������� ���',
    '��������', '��������� (���.)', '�������� (���.)', '���������� �-�', '���������� �����', '��������', '�������������', '������', '�����������',
    '������� ���� ���', '��������', '�����������', '���������', '����������', '����-��������', '��������� �����', '��������', '������������� ��',
    '����������', '���������� ����', '������������', '���. ����� ����', '���� ������ ��', '�������', '��������', '���������', '���������', '���������',
    '�������� ���', '����� �����.���', '������������ ���', '��������� ���', '���������� ���', '������� ���', '�������� ���', '��������� ���',
    '�������� ���', '�-� ��������.���', '��������� ���', '��������� ���', '�����. ���� ���', '���������� ���', '��������� ���', '�������� ���',
    '���������. ���', '��� ���', '�������� ���', '��������� ���', '�����. ��� ���', '������������ ���', '����������-� ���', '�����. ����� ���',
    '�������� ���', '�����������. ���', '������� ���', '����������� ���', '������-���� ���', '����� ���', '��. �������� ���', '���������', '������������',
    '���������-�����.', '�������', '���������-�� �-�', '�������', '�����������.��-�', '���� ������ ����', '��������� ���� �', '������������� 2',
     '������ 2', '���������� ����', '���. ����� ���', '�������', '��������', '�����������', '����', '��������� ���� �', '�����������',
     '������� ��������', '��������', '��������(�����.)', '�������-�(�����)', '����(�����.)', '����.����(�����)', '�������.��-�����', '������',
     '��������', '��������', '��������� �����', '���������������', '�����������', '�����������', '���������� ���', '������� ���', '��������', '��������',
     '����������', '������', '��.�������������', '������������', '����������', '���-���������', '������', '������������', '�������������',
     '������������ ���', '���������', '��� ���������', '������������ ���', '��������� ���.', '��������� ���', '����������-� ���', '�������� �����-�',
                        '�������', '����������-� ���', '�����������(��)', '������',
    '������������', '�����������', '��-� ������. ���', '�������.��-� ���', '�����������', '���������', '���������� ���', '��������', '������� ���� ���',
    '������� ���', '���������� ���', '������������ ���', '��������� ���', '����������� �-�', '�������� �������',
    '���������-� ���', '��������� (���)', '���������-� ���', '�', '�������', '�������� �������', '���������', '���������', '������', '������������ ��']
    
    with GigaChat(credentials='MmQyNzJmNzAtNTRhMi00ZTg5LTgxNGMtMWQ2OTJhNzg0MmFmOmU5MTY0YjlmLTVhODMtNGI2Yi1iZTU5LTJkYzM5ZjJiNjk3YQ==', verify_ssl_certs=False) as giga:
        #response_text = ', '.join(map(str,candidate_labels)) + '������ �� ������� �� ������, ������� ����������� � ��������� �����������' + input_text
        response_text = '������ �� ������� ����� ������, ������� ����������� � ��������� �����������, ���� ������ �������� ������� � ������� "<�������� �������>": ' + input_text
        response = giga.chat(response_text)
        date2 = response.choices[0].message.content
   
    print(date2)
    l = list(map(similar_text,[date2.lower() for i in range(len(candidate_labels))],candidate_labels))
    res = list()
    for i in range(len(candidate_labels)):
      res.append([l[i],candidate_labels[i]])
    res1 = sorted(res,reverse=True)[0][1]

    try:
        res = bool(parser.parse(date1))
    except ValueError:
        try:
            date1 = extract_relative_date_from_text(date1).date().strftime('%d.%m.%Y')
        except:
            date1="03.04.2024"

    return (date1,res1)


import openpyxl
from datetime import datetime

def is_date_column(sheet, col_num, dates):
    cell_value = sheet.cell(row=1, column=col_num).value
    if isinstance(cell_value, datetime):
        date_str = cell_value.strftime('%d.%m.%Y')
        if date_str in dates:
            return True
    return False

def get_sum_for_station_and_dates(file_path, station_name, dates):
    # ��������� ���� Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # �������� ������� �������� ��� ������� � ���
    station_column = None
    date_columns = []
    for col_num in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col_num).value == "�������":
            station_column = col_num
        elif is_date_column(sheet, col_num, dates):
            date_columns.append(col_num)

    if station_column is None:
        return "������� '�������' �� ������ � �����"
    
    if not date_columns:
        return "�� ���a �� ��� ��������� � ������� �� ������� � �����"

    # ���� ������, ��������������� ��������� �������
    station_rows = []
    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=station_column).value == station_name:
            station_rows.append(row_num)

    if not station_rows:
        return f"������� '{station_name}' �� ������� � �����"

    # ��������� �������� ��� ������ ����
    total_sum = 0
    for row_num in station_rows:
        for date_column in date_columns:
            value = sheet.cell(row=row_num, column=date_column).value
            if value is not None:
                total_sum += value

    return total_sum

# ������ �������������
file_path = "./test.xlsx"
#station_name = "�.�������������"
#dates = ["01.04.2024", "02.04.2024", "03.04.2024"]

#result = get_sum_for_station_and_dates(file_path, station_name, dates)
#print("����� �������� ��� ��������� ������� � ���:", result)


import speech_recognition as sr
from pydub import AudioSegment
import os

def recognize_speech(file_path):
    # ����������� ���� OGG � WAV
    sound = AudioSegment.from_ogg(file_path)
    wav_path = "temp.wav"
    sound.export(wav_path, format="wav")

    recognizer = sr.Recognizer()

    # ��������� WAV ����
    with sr.AudioFile(wav_path) as source:
        audio_data = recognizer.record(source)  # ��������� �����������

    try:
        # �������� ������������� ����
        text = recognizer.recognize_google(audio_data, language="ru-RU")
        return text
    except sr.UnknownValueError:
        return "�� ������� ���������� ����"
    except sr.RequestError as e:
        return f"������ ��� �������� ������� � ������� �������������: {e}"
    finally:
        # ������� ��������� WAV ����
        os.remove(wav_path)


bot = telebot.TeleBot('6983581621:AAF7YZoFUBaEhM-t2HVsg1qetD9Mn95ppOk')

@bot.message_handler()
def save_message(message):
    #with open("messages.txt", "a", encoding="utf-8") as file:
        #file.write(message.text + "\n")
    print(f'������: {message.text}')
    res=classify(message.text)
    print(f'�����: {res}')
    bot.reply_to(message, f'������������ ������: {res}')
    result = get_sum_for_station_and_dates(file_path, res[1], res[0])
    bot.reply_to(message, f'���������: {result}')
    print(f'�����: {result}')

@bot.message_handler(content_types=['voice'])
def voice_processing(message):
    file_info = bot.get_file(message.voice.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    with open('new_file.ogg', 'wb+') as new_file:
        new_file.write(downloaded_file)
    recognized_text = recognize_speech("./new_file.ogg")
    print("������������ �����: ", recognized_text)
    res=classify(recognized_text)
    print(f'�����: {res}')
    bot.reply_to(message, f'������������ ������: {res}')
    result = get_sum_for_station_and_dates(file_path, res[1], res[0])
    bot.reply_to(message, f'���������: {result}')
    print(f'�����: {result}')
   
bot.polling(none_stop=True)

